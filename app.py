from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
import xml.etree.ElementTree as ET
from urllib.parse import urlparse
import time
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import logging
import warnings
import io
import os
from datetime import datetime

warnings.filterwarnings('ignore')

app = Flask(__name__)
CORS(app)

# Configuration
MAX_WORKERS = 12
REQUEST_TIMEOUT = 15
RETRY_ATTEMPTS = 2

# Enhanced categorization with proper naming
CATEGORIES = [
    ('Products', '/products/'),
    ('Product', '/product/'),
    ('Product Categories', '/product-category/'),
    ('Collections', '/collections/'),
    ('Blog', '/blog/'),
    ('Posts', '/posts/'),
    ('Blogs', '/blogs/'),
    ('Categories', '/categories/'),
    ('Category', '/category/'),
    ('Portfolio', '/portfolio/'),
    ('Events', '/events/'),
    ('Archive', '/archive/'),
    ('Tags', '/tags/'),
    ('Tag', '/tag/'),
    ('Author', '/author/'),
    ('Users', '/users/'),
    ('Gallery', '/gallery/'),
    ('Downloads', '/downloads/'),
    ('Docs', '/docs/'),
    ('Documentation', '/documentation/'),
    ('Testimonials', '/testimonials/'),
    ('Testimonial', '/testimonial/'),
    ('Node', '/node/'),
    ('Content', '/content/'),
    ('Article', '/article/'),
    ('Articles', '/articles/'),
    ('Page', '/page/'),
    ('Pages', '/pages/'),
    ('Projects', '/projects/'),
    ('Case Studies', '/case-studies/'),
    ('Locations', '/locations/'),
    ('Service', '/service/'),
    ('Services', '/services/'),
    ('Courses', '/courses/'),
    ('Jobs', '/jobs/'),
    ('Published', '/published/'),
    ('Draft', '/draft/'),
    ('Departments', '/departments/'),
    ('Videos', '/videos/'),
    ('Video', '/video/'),
    ('Images', '/images/'),
    ('Image', '/image/'),
    ('WP Content', '/wp-content/'),
    ('Tools', '/tools/'),
    ('Tool', '/tool/'),
    ('Help', '/help/'),
    ('FAQ', '/faq/'),
    ('Helps', '/helps/'),
    ('Teams', '/teams/'),
    ('Team', '/team/'),
    ('Team Members', '/team-members/'),
    ('Team Member', '/team-member/'),
    ('Member', '/member/'),
    ('Resources', '/resources/'),
    ('News', '/news/')
]

# Setup logging
logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)

def create_session():
    """Create reusable session with connection pooling"""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept-Language': 'en-US,en;q=0.5'
    })
    return session

def extract_metadata(soup, url):
    """Extract SEO elements with comprehensive duplicate detection and reporting"""
    try:
        # Enhanced title extraction - check for multiple titles
        head = soup.find('head')
        titles = []
        title_issues = []

        if head:
            title_tags = head.find_all('title')  # Get ALL title tags
            titles = [tag.text.strip() for tag in title_tags if tag.text and tag.text.strip()]

            if len(titles) > 1:
                title_issues.append(f"⚠️ MULTIPLE TITLES FOUND ({len(titles)})")

        # Format title output
        if titles:
            if len(titles) == 1:
                final_title = titles[0]
            else:
                # Show all titles with numbering for multiple instances
                numbered_titles = [f"{i+1}. {title}" for i, title in enumerate(titles)]
                final_title = "\n".join(numbered_titles)
                if title_issues:
                    final_title = f"{title_issues[0]}\n{final_title}"
        else:
            final_title = ""  # Keep empty instead of error message

        # Enhanced meta description extraction - check for multiple descriptions
        meta_descriptions = []
        desc_issues = []

        # Get all meta description tags (including variations)
        desc_tags = soup.find_all('meta', {'name': ['description', 'Description', 'DESCRIPTION']})
        meta_descriptions = [tag.get('content', '').strip() for tag in desc_tags if tag.get('content', '').strip()]

        if len(meta_descriptions) > 1:
            desc_issues.append(f"⚠️ MULTIPLE META DESCRIPTIONS ({len(meta_descriptions)})")

        # Format description output
        if meta_descriptions:
            if len(meta_descriptions) == 1:
                final_description = meta_descriptions[0]
            else:
                # Show all descriptions with numbering
                numbered_descriptions = [f"{i+1}. {desc}" for i, desc in enumerate(meta_descriptions)]
                final_description = "\n".join(numbered_descriptions)
                if desc_issues:
                    final_description = f"{desc_issues[0]}\n{final_description}"
        else:
            final_description = ""  # Keep empty instead of error message

        # Enhanced H1 extraction with issue detection
        h1_tags = soup.find_all('h1')
        h1_texts = [h1.text.strip() for h1 in h1_tags if h1.text and h1.text.strip()]
        h1_issues = []

        if len(h1_texts) > 1:
            h1_issues.append(f"⚠️ MULTIPLE H1 TAGS ({len(h1_texts)})")

        # Format H1 output
        if h1_texts:
            if len(h1_texts) == 1:
                final_h1 = h1_texts[0]
            else:
                numbered_h1s = [f"{i+1}. {h1}" for i, h1 in enumerate(h1_texts)]
                final_h1 = "\n".join(numbered_h1s)
                if h1_issues:
                    final_h1 = f"{h1_issues[0]}\n{final_h1}"
        else:
            final_h1 = ""  # Keep empty instead of error message

        # Enhanced canonical URL extraction - check for multiple canonical tags
        canonical_tags = soup.find_all('link', {'rel': 'canonical'})
        canonicals = [tag.get('href', '').strip() for tag in canonical_tags if tag.get('href', '').strip()]

        if len(canonicals) > 1:
            final_canonical = f"⚠️ MULTIPLE CANONICAL TAGS ({len(canonicals)})\n" + "\n".join([f"{i+1}. {can}" for i, can in enumerate(canonicals)])
        elif len(canonicals) == 1:
            final_canonical = canonicals[0]
        else:
            final_canonical = ""

        # Enhanced robots tag extraction
        robots_tags = soup.find_all('meta', {'name': 'robots'})
        robots_contents = [tag.get('content', '').lower() for tag in robots_tags if tag.get('content')]

        # Check for noindex in any robots tag
        meta_robots_noindex = any('noindex' in content for content in robots_contents)

        # Report multiple robots tags
        robots_display = 'Yes' if meta_robots_noindex else 'No'
        if len(robots_tags) > 1:
            robots_display = f"⚠️ MULTIPLE ROBOTS TAGS - {robots_display}"

        return {
            'Meta Title': final_title,
            'Meta Description': final_description,
            'H1': final_h1,
            'Canonical URL': final_canonical,
            'Meta Robots Noindex': robots_display
        }

    except Exception as e:
        logging.error(f"Metadata extraction error for {url}: {e}")
        return {
            'Meta Title': f'❌ EXTRACTION ERROR: {str(e)}',
            'Meta Description': f'❌ EXTRACTION ERROR: {str(e)}',
            'H1': f'❌ EXTRACTION ERROR: {str(e)}',
            'Canonical URL': f'❌ EXTRACTION ERROR: {str(e)}',
            'Meta Robots Noindex': 'Error'
        }

def apply_excel_optimizations(worksheet):
    """Enhanced Excel formatting with conditional formatting for SEO issues"""
    column_dimensions = {
        'A': 45,  # Original URL
        'B': 45,  # Final URL
        'C': 50,  # Meta Title (increased for issue warnings)
        'D': 70,  # Meta Description (increased for issue warnings)
        'E': 50,  # H1 (increased for issue warnings)
        'F': 15,  # Status Code
        'G': 15,  # Redirect Count
        'H': 45,  # Canonical URL
        'I': 25   # Meta Robots Noindex (increased for warnings)
    }

    header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    # Colors for different issue types
    warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow

    for col, width in column_dimensions.items():
        worksheet.column_dimensions[col].width = width
        header_cell = worksheet[f'{col}1']
        header_cell.fill = header_fill
        header_cell.font = header_font

    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment
            cell_value = str(cell.value) if cell.value else ''

            # URL formatting (columns A, B, H)
            if cell.column_letter in ['A', 'B', 'H'] and not ('❌' in cell_value or '⚠️' in cell_value):
                cell.font = Font(color='0000FF', underline='single')
                cell.hyperlink = cell.value

            # Only highlight extraction errors, not missing values
            elif '❌' in cell_value and 'EXTRACTION ERROR' in cell_value:  # Only extraction errors
                cell.font = Font(color='CC0000', bold=True)
            elif '⚠️' in cell_value:  # Warnings for multiple items
                cell.fill = warning_fill
                cell.font = Font(color='FF6600', bold=True)

            # Meta Robots column special formatting
            elif cell.column_letter == 'I':
                if 'Yes' in cell_value:
                    cell.font = Font(color='FF0000', bold=True)
                elif 'No' in cell_value:
                    cell.font = Font(color='008000', bold=True)
                elif '⚠️' in cell_value:
                    cell.fill = warning_fill
                    cell.font = Font(color='FF6600', bold=True)

            # Multiple items formatting (numbered lists)
            elif '\n1. ' in cell_value and not ('❌' in cell_value or '⚠️' in cell_value):
                cell.font = Font(color='2E75B6', bold=True)

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = worksheet.dimensions

def process_url(session, url):
    """Process URL with enhanced error handling"""
    final_url = url
    status_code = None
    redirect_count = 0
    metadata = {}

    for attempt in range(RETRY_ATTEMPTS):
        try:
            response = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            status_code = response.status_code
            final_url = response.url
            redirect_count = len(response.history)

            soup = BeautifulSoup(response.content, 'html.parser')  # Changed from lxml
            metadata = extract_metadata(soup, final_url)
            break

        except Exception as e:
            logging.warning(f"Attempt {attempt+1} failed for {url}: {str(e)}")
            if attempt == RETRY_ATTEMPTS -1:
                status_code = 'Error'

    return {
        'Original URL': url,
        'Final URL': final_url,
        'Meta Title': metadata.get('Meta Title', ''),
        'Meta Description': metadata.get('Meta Description', ''),
        'H1': metadata.get('H1', ''),
        'Status Code': status_code,
        'Redirect Count': redirect_count,
        'Canonical URL': metadata.get('Canonical URL', ''),
        'Meta Robots Noindex': metadata.get('Meta Robots Noindex', 'No')
    }

def get_sitemap_urls(session, sitemap_url):
    """Fetch and parse sitemap hierarchy"""
    try:
        response = session.get(sitemap_url, timeout=REQUEST_TIMEOUT)
        root = ET.fromstring(response.content)
        namespaces = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
        urls = []

        if root.tag.endswith('sitemapindex'):
            for sitemap in root.findall('sm:sitemap', namespaces):
                child_url = sitemap.find('sm:loc', namespaces).text
                urls += get_sitemap_urls(session, child_url)
        else:
            for url in root.findall('sm:url', namespaces):
                loc = url.find('sm:loc', namespaces)
                if loc is not None:
                    urls.append(loc.text)
        return list(set(urls))
    except Exception as e:
        logging.error(f"Sitemap processing error: {e}")
        return []

def categorize_data(metadata_list):
    """Categorize URLs based on Original URL with enhanced matching"""
    # Initialize categories with proper names
    categorized = {category[0]: [] for category in CATEGORIES}
    categorized['Main'] = []

    for data in metadata_list:
        parsed = urlparse(data['Original URL'])
        matched = False

        # Sort categories by pattern length (longest first) for better matching
        sorted_categories = sorted(CATEGORIES, key=lambda x: len(x[1]), reverse=True)

        for name, pattern in sorted_categories:
            if pattern in parsed.path:
                categorized[name].append(data)
                matched = True
                break

        if not matched:
            categorized['Main'].append(data)

    return categorized

def create_excel_report(categorized_data, filename):
    """Create Excel file with structured data and proper sheet naming"""
    columns_order = [
        'Original URL',
        'Final URL',
        'Meta Title',
        'Meta Description',
        'H1',
        'Status Code',
        'Redirect Count',
        'Canonical URL',
        'Meta Robots Noindex'
    ]

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create Main sheet first
    if categorized_data['Main']:
        ws = wb.create_sheet(title='Main')
        
        # Add headers
        for col_num, header in enumerate(columns_order, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, item in enumerate(categorized_data['Main'], 2):
            for col_num, header in enumerate(columns_order, 1):
                ws.cell(row=row_num, column=col_num, value=item.get(header, ''))
        
        apply_excel_optimizations(ws)

    # Create category sheets with proper naming
    for category_name, _ in CATEGORIES:
        if categorized_data[category_name]:
            sheet_name = category_name[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            
            # Add headers
            for col_num, header in enumerate(columns_order, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Add data
            for row_num, item in enumerate(categorized_data[category_name], 2):
                for col_num, header in enumerate(columns_order, 1):
                    ws.cell(row=row_num, column=col_num, value=item.get(header, ''))
            
            apply_excel_optimizations(ws)

    # Handle empty sitemap case
    if all(not v for v in categorized_data.values()):
        ws = wb.create_sheet(title='Info')
        ws.cell(row=1, column=1, value='No URLs found in sitemap')

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/')
def home():
    """Basic API info"""
    return jsonify({
        'message': 'SEO Analyzer Pro API - Advanced Version',
        'status': 'running',
        'version': '2.0',
        'endpoints': ['/api/analyze', '/api/download/<filename>']
    })

@app.route('/api/analyze', methods=['POST'])
def analyze_sitemap():
    """Main analysis endpoint with advanced features"""
    try:
        data = request.get_json()
        sitemap_url = data.get('sitemap_url', '').strip()
        
        if not sitemap_url:
            return jsonify({'error': 'Sitemap URL required'}), 400

        session = create_session()
        start_time = time.time()
        
        logging.info("🚀 Starting advanced sitemap processing")
        
        # Get URLs from sitemap
        all_urls = get_sitemap_urls(session, sitemap_url)
        logging.info(f"🌐 Found {len(all_urls)} unique URLs")
        
        if not all_urls:
            return jsonify({'error': 'No URLs found in sitemap'}), 400

        # Process URLs with advanced analysis
        metadata_list = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_url, session, url): url for url in all_urls}
            
            for future in as_completed(futures):
                result = future.result()
                metadata_list.append(result)

        # Categorize and analyze results with enhanced logic
        categorized = categorize_data(metadata_list)
        analysis_time = time.time() - start_time
        
        # Advanced issue counting
        issues = {
            'multipleTitles': sum(1 for item in metadata_list if '⚠️ MULTIPLE TITLES' in str(item.get('Meta Title', ''))),
            'missingDescriptions': sum(1 for item in metadata_list if not item.get('Meta Description', '').strip()),
            'multipleH1s': sum(1 for item in metadata_list if '⚠️ MULTIPLE H1' in str(item.get('H1', ''))),
            'noindexPages': sum(1 for item in metadata_list if 'Yes' in str(item.get('Meta Robots Noindex', ''))),
            'longTitles': sum(1 for item in metadata_list if len(str(item.get('Meta Title', ''))) > 60),
            'shortDescriptions': sum(1 for item in metadata_list if 0 < len(str(item.get('Meta Description', ''))) < 120),
            'errors': sum(1 for item in metadata_list if 'Error' in str(item.get('Status Code', '')))
        }
        
        # Generate filename
        domain = urlparse(sitemap_url).netloc.replace('.', '_')
        filename = f"{domain}_advanced_seo_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create Excel file with advanced formatting
        excel_data = create_excel_report(categorized, filename)
        
        # Store for download
        if not hasattr(app, 'reports'):
            app.reports = {}
        app.reports[filename] = excel_data
        
        # Count categories
        categories = {name: len(data) for name, data in categorized.items() if data}
        
        session.close()
        logging.info(f"⏱ Advanced analysis completed in {analysis_time:.2f} seconds")

        return jsonify({
            'success': True,
            'totalUrls': len(metadata_list),
            'categories': categories,
            'issues': issues,
            'analysisTime': f"{analysis_time:.1f} seconds",
            'downloadFilename': filename,
            'stats': {
                'processed': len(metadata_list),
                'errors': issues['errors'],
                'warnings': sum([issues['multipleTitles'], issues['multipleH1s'], issues['longTitles'], issues['shortDescriptions']]),
                'healthy': len(metadata_list) - issues['errors'] - sum([issues['multipleTitles'], issues['multipleH1s']])
            }
        })
        
    except Exception as e:
        logging.error(f"Advanced analysis error: {str(e)}")
        return jsonify({'error': f'Analysis failed: {str(e)}'}), 500

@app.route('/api/download/<filename>')
def download_report(filename):
    """Download Excel report"""
    try:
        if not hasattr(app, 'reports') or filename not in app.reports:
            return jsonify({'error': 'Report not found'}), 404
        
        excel_data = app.reports[filename]
        excel_data.seek(0)
        
        return send_file(
            excel_data,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    print("🚀 Starting SEO Analyzer Pro - Advanced Version...")
    app.run(debug=False, host='0.0.0.0', port=port)
